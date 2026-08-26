from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .detect import ColumnMapping, detect_columns
from .models import InputCase
from .utils import clean_text


def read_base_auto(path: Path, mapping: ColumnMapping | None = None) -> tuple[list[InputCase], ColumnMapping]:
    """Lee CUALQUIER Excel usando deteccion automatica de columnas (o un mapping dado).

    A diferencia de read_base (que exige la hoja/columnas fijas del flujo original),
    esta version localiza sola la hoja y las columnas de radicado/demandante/demandado,
    para el aplicativo web donde el usuario sube bases con estructuras distintas.
    """
    if mapping is None:
        mapping = detect_columns(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[mapping.sheet] if mapping.sheet in wb.sheetnames else wb.active
    cases: list[InputCase] = []
    seen: set[str] = set()
    start_row = mapping.header_row + 1 if mapping.header_row else 1
    for row_number, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        radicado = clean_text(_cell(row, mapping.radicado_col))
        if not radicado or radicado in seen:
            continue
        seen.add(radicado)
        input_id = clean_text(_cell(row, mapping.id_col)) if mapping.id_col is not None else ""
        if not input_id:
            input_id = str(len(cases) + 1)
        cases.append(
            InputCase(
                input_id=input_id,
                radicado=radicado,
                demandante=clean_text(_cell(row, mapping.demandante_col)),
                demandado=clean_text(_cell(row, mapping.demandado_col)),
                excel_row=row_number,
            )
        )
    wb.close()
    return cases, mapping


def _cell(row: tuple, col: int | None) -> object:
    if col is None or col >= len(row):
        return None
    return row[col]


def read_base(path: Path) -> list[InputCase]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["ULTIMOS MOVIMIENTOS"] if "ULTIMOS MOVIMIENTOS" in wb.sheetnames else wb.active
    headers = [clean_text(cell.value).upper() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    required = {"ID", "RADICADO", "DEMANDANTE", "DEMANDADO"}
    missing = required.difference(headers)
    if missing:
        raise ValueError(f"Faltan columnas requeridas en la base: {', '.join(sorted(missing))}")
    indexes = {name: headers.index(name) for name in required}
    cases: list[InputCase] = []
    seen: set[str] = set()
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        radicado = clean_text(row[indexes["RADICADO"]])
        if not radicado or radicado in seen:
            continue
        seen.add(radicado)
        cases.append(
            InputCase(
                input_id=clean_text(row[indexes["ID"]]),
                radicado=radicado,
                demandante=clean_text(row[indexes["DEMANDANTE"]]),
                demandado=clean_text(row[indexes["DEMANDADO"]]),
                excel_row=row_number,
            )
        )
    return cases
