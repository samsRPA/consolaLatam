"""Deteccion automatica de columnas en una base de Excel arbitraria.

El scraper original exigia una hoja "ULTIMOS MOVIMIENTOS" con columnas fijas
(ID, RADICADO, DEMANDANTE, DEMANDADO). Este modulo permite recibir CUALQUIER Excel:
localiza la hoja y las columnas correctas por contenido (regex de radicado) y por
nombre de encabezado (tolerante a acentos, mayusculas y sinonimos), para que el
usuario no tenga que reordenar ni renombrar nada antes de consultar.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from .utils import clean_text, normalize_for_match

# Regex laxo (a proposito) para RECONOCER radicados dentro de una celda cualquiera.
# Es mas permisivo que utils.RADICADO_RE porque aqui solo queremos "puntuar" columnas,
# no validar estrictamente: acepta espacios alrededor y hasta 5/4/2 digitos por bloque.
RADICADO_SCAN_RE = re.compile(
    r"^\s*\d{1,5}-\d{4}-\d{1,4}-\d{4}-[A-Za-z]{2}-[A-Za-z]{2}-\d{1,2}\s*$"
)

# Sinonimos que aparecen en bases judiciales peruanas reales para cada parte.
DEMANDANTE_HINTS = ("DEMANDANTE", "ACTOR", "RECURRENTE", "DENUNCIANTE", "SOLICITANTE", "PARTE ACTIVA")
DEMANDADO_HINTS = ("DEMANDADO", "EMPLAZADO", "DENUNCIADO", "REQUERIDO", "PARTE PASIVA")
ID_HINTS = ("ID", "ITEM", "N°", "NRO", "NUMERO", "#", "CODIGO", "COD")


@dataclass
class ColumnMapping:
    """Resultado de la deteccion. Los indices son 0-based sobre la fila de la hoja."""

    sheet: str
    header_row: int  # fila (1-based) donde estan los encabezados; 0 = no hay encabezados
    radicado_col: int
    demandante_col: int | None
    demandado_col: int | None
    id_col: int | None
    confidence: float
    preview: list[dict[str, str]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def to_json(self) -> dict[str, object]:
        return {
            "sheet": self.sheet,
            "header_row": self.header_row,
            "radicado_col": self.radicado_col,
            "demandante_col": self.demandante_col,
            "demandado_col": self.demandado_col,
            "id_col": self.id_col,
            "confidence": round(self.confidence, 3),
            "preview": self.preview,
            "warnings": self.warnings,
        }

    @classmethod
    def from_json(cls, data: dict[str, object]) -> "ColumnMapping":
        return cls(
            sheet=str(data["sheet"]),
            header_row=int(data["header_row"]),
            radicado_col=int(data["radicado_col"]),
            demandante_col=_opt_int(data.get("demandante_col")),
            demandado_col=_opt_int(data.get("demandado_col")),
            id_col=_opt_int(data.get("id_col")),
            confidence=float(data.get("confidence", 0.0)),
            preview=list(data.get("preview", [])),  # type: ignore[arg-type]
            warnings=list(data.get("warnings", [])),  # type: ignore[arg-type]
        )


def _opt_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    return int(value)  # type: ignore[arg-type]


class DetectionError(ValueError):
    """No se pudo identificar una columna de radicados en ninguna hoja."""


def _radicado_score(cells: list[str]) -> tuple[int, int]:
    """Devuelve (coincidencias, celdas_no_vacias) del regex de radicado en una columna."""
    matches = 0
    non_empty = 0
    for cell in cells:
        text = clean_text(cell)
        if not text:
            continue
        non_empty += 1
        if RADICADO_SCAN_RE.match(text):
            matches += 1
    return matches, non_empty


def _match_header(header: str, hints: tuple[str, ...]) -> bool:
    norm = normalize_for_match(header)
    return any(hint in norm for hint in hints)


def detect_columns(path: Path) -> ColumnMapping:
    """Analiza TODAS las hojas del workbook y elige la que contiene radicados.

    Estrategia:
      1. Por cada hoja, por cada columna, cuenta cuantas celdas hacen match de radicado.
      2. Gana la columna/hoja con mayor ratio de coincidencias (min 40% de celdas).
      3. Detecta la fila de encabezados (la primera fila SIN radicados, arriba de los datos).
      4. Asigna demandante/demandado por nombre de encabezado; si faltan, deja None.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    best: ColumnMapping | None = None
    best_ratio = 0.0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            continue
        n_cols = max((len(r) for r in rows), default=0)
        if n_cols == 0:
            continue

        # Puntua cada columna como candidata a radicado, mirando SOLO de la fila 2 en
        # adelante (la 1 suele ser encabezado). Si no hay encabezado, la fila 1 tambien
        # hara match y no se pierde nada.
        for col in range(n_cols):
            column_cells = [row[col] if col < len(row) else None for row in rows[1:]]
            matches, non_empty = _radicado_score([str(c) if c is not None else "" for c in column_cells])
            if non_empty == 0:
                continue
            ratio = matches / non_empty
            if ratio > best_ratio and ratio >= 0.4 and matches >= 1:
                best_ratio = ratio
                best = _build_mapping(sheet_name, rows, col, n_cols, ratio)

    wb.close()
    if best is None:
        raise DetectionError(
            "No se encontro ninguna columna con radicados (formato NNNNN-AAAA-N-DDDD-XX-XX-NN) "
            "en ninguna hoja del archivo."
        )
    return best


def _build_mapping(sheet_name: str, rows: list[list], radicado_col: int, n_cols: int, ratio: float) -> ColumnMapping:
    warnings: list[str] = []

    # Fila de encabezado: la fila 1 se considera encabezado si su celda en la columna de
    # radicado NO es un radicado (es texto tipo "RADICADO"). Si la fila 1 ya es un
    # radicado, no hay encabezados.
    first_cell = clean_text(rows[0][radicado_col]) if radicado_col < len(rows[0]) else ""
    has_header = not RADICADO_SCAN_RE.match(first_cell)
    header_row = 1 if has_header else 0
    headers = [clean_text(c) for c in rows[0]] if has_header else ["" for _ in range(n_cols)]

    demandante_col = _find_header_col(headers, DEMANDANTE_HINTS)
    demandado_col = _find_header_col(headers, DEMANDADO_HINTS)
    id_col = _find_header_col(headers, ID_HINTS)

    # Fallback posicional cuando no hay encabezados nombrados: en las bases tipicas el
    # orden es ...RADICADO | DEMANDANTE | DEMANDADO..., asi que asumimos las dos columnas
    # inmediatamente a la derecha del radicado.
    if demandante_col is None and not has_header:
        cand = radicado_col + 1
        if cand < n_cols:
            demandante_col = cand
            warnings.append("Sin encabezados: se asumio DEMANDANTE en la columna a la derecha del radicado.")
    if demandado_col is None and not has_header:
        cand = radicado_col + 2
        if cand < n_cols:
            demandado_col = cand
            warnings.append("Sin encabezados: se asumio DEMANDADO dos columnas a la derecha del radicado.")

    if demandante_col is None:
        warnings.append("No se detecto columna DEMANDANTE; se usara solo el DEMANDADO como parte de busqueda.")
    if demandado_col is None:
        warnings.append("No se detecto columna DEMANDADO; se usara solo el DEMANDANTE como parte de busqueda.")

    data_start = 1 if has_header else 0
    preview: list[dict[str, str]] = []
    for row in rows[data_start : data_start + 8]:
        rad = clean_text(row[radicado_col]) if radicado_col < len(row) else ""
        if not rad:
            continue
        preview.append(
            {
                "radicado": rad,
                "demandante": clean_text(row[demandante_col]) if demandante_col is not None and demandante_col < len(row) else "",
                "demandado": clean_text(row[demandado_col]) if demandado_col is not None and demandado_col < len(row) else "",
            }
        )

    return ColumnMapping(
        sheet=sheet_name,
        header_row=header_row,
        radicado_col=radicado_col,
        demandante_col=demandante_col,
        demandado_col=demandado_col,
        id_col=id_col,
        confidence=ratio,
        preview=preview,
        warnings=warnings,
    )


def _find_header_col(headers: list[str], hints: tuple[str, ...]) -> int | None:
    for index, header in enumerate(headers):
        if header and _match_header(header, hints):
            return index
    return None
