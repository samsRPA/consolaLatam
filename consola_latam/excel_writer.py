from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .utils import clean_text


# Colores del modo claro del aplicativo web.
BRAND_FOREST = "9A1413"
BRAND_MIST = "F5F5F5"


def _style_header(row, color: str = BRAND_FOREST) -> None:
    fill = PatternFill("solid", fgColor=color)
    for cell in row:
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _cell_value(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _append_row(ws, values) -> None:
    ws.append([_cell_value(value) for value in values])


# ============================================================================
#  Ecuador: el bot devuelve datos planos del proceso (sin demandante/demandado ni
#  historial de actuaciones), asi que el Excel es distinto al de Peru: una sola hoja
#  resumen con UNA COLUMNA POR CAMPO (nada de hojas de detalle por radicado, que
#  quedarian vacias). Tonos azules para que combine con el tema de esa sede.
# ============================================================================

BRAND_EC_BLUE = "13398A"

ECUADOR_HEADERS = [
    "RADICADO", "CLIENTE", "MATERIA", "TIPO DE ACCIÓN", "DELITO/ASUNTO",
    "JUDICATURA", "ID JUDICATURA", "CIUDAD", "FECHA DE INGRESO", "ACTORES", "DEMANDADOS",
]
ECUADOR_COL_WIDTHS = [18, 22, 12, 16, 34, 46, 14, 14, 18, 34, 34]


def write_ecuador_processes_workbook(processes: list[dict], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Procesos Ecuador"
    _append_row(ws, ECUADOR_HEADERS)
    _style_header(ws[1], BRAND_EC_BLUE)
    for proc in processes:
        detail = proc.get("detail") or {}
        reporte = detail.get("reporte") or {}
        radicado = proc.get("radicado", "")
        cliente = proc.get("client_name", "")
        materia = reporte.get("Materia") or proc.get("materia", "")
        tipo_accion = reporte.get("Tipo de Acción") or proc.get("estado", "")
        delito_asunto = reporte.get("Delito/Asunto", "")
        judicatura = reporte.get("Judicatura") or proc.get("organo", "")
        ciudad = reporte.get("Ciudad", "")
        fecha_ingreso = reporte.get("Fecha de Ingreso", "")
        # Un radicado puede traer varios expedientes (uno por idJudicatura), cada uno con su
        # propia judicatura/ciudad; se repite la fila con los datos comunes del radicado y
        # cambian idJudicatura/judicatura/ciudad/actores/demandados. Si el expediente no trae
        # judicatura/ciudad propia (formato viejo) se usa la del radicado como respaldo.
        expedientes = detail.get("expedientes") or [{}]
        for exp in expedientes:
            _append_row(ws, [
                radicado, cliente, materia, tipo_accion, delito_asunto,
                exp.get("nombreJudicatura") or judicatura, exp.get("idJudicatura", ""),
                exp.get("ciudad") or ciudad, fecha_ingreso,
                "; ".join(exp.get("actores") or []),
                "; ".join(exp.get("demandados") or []),
            ])
    for idx, width in enumerate(ECUADOR_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "top" if cell.row > 1 else "center"
            cell.alignment = alignment
            if cell.row > 1:
                font = copy(cell.font)
                font.name = "Calibri"
                cell.font = font
                if cell.row % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="EAF0FB")
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ============================================================================
#  Peru (bot externo): igual que Ecuador, el bot devuelve datos planos del proceso
#  (courtOfficeCode, caseReport, actorsRama) SIN historial de actuaciones. Una sola hoja
#  resumen con una columna por dato, tonos rojos (los de siempre en Peru).
# ============================================================================

PERU_BOT_HEADERS = [
    "RADICADO", "CLIENTE", "DESPACHO", "ESPECIALISTA LEGAL", "FECHA INICIO", "MATERIA",
    "ETAPA PROCESAL", "UBICACION", "ESPECIALIDAD", "ESTADO", "DISTRITO JUDICIAL",
    "DEMANDANTES", "DEMANDADOS", "VALOR PARTE",
    "TIPO DOCUMENTO", "NRO DOCUMENTO", "CODIGO", "FECHA EMISION", "FECHA NACIMIENTO",
    "ERROR",
]
PERU_BOT_REPORT_KEYS = [
    "ESPECIALISTA LEGAL", "FECHA INICIO", "MATERIA", "ETAPA PROCESAL",
    "UBICACION", "ESPECIALIDAD", "ESTADO", "DISTRITO JUDICIAL",
]
PERU_BOT_DOCUMENTO_KEYS = ["tipoDocumento", "numeroDocumento", "codigo", "fechaEmision", "fechaNacimiento"]
PERU_BOT_COL_WIDTHS = [18, 22, 26, 26, 14, 26, 14, 20, 16, 22, 18, 34, 34, 22, 16, 18, 12, 16, 18, 34]


def _grouped_actor_names(detail: dict, tipo: str, fallback: str) -> str:
    """Demandante(s)/demandado(s) son siempre una lista (actorsRama puede traer varios
    actores del mismo tipo): se agrupa desde ahi, uno por linea dentro de la misma celda,
    en vez del string ya unido con ';' que trae el proceso (que solo se usa de respaldo
    si el proceso no tiene actorsRama, ej. un registro de error)."""
    names = []
    for actor in detail.get("actorsRama") or []:
        if clean_text(actor.get("tipo_sujeto")).upper() != tipo:
            continue
        nombre = clean_text(actor.get("nombre_actor"))
        if nombre:
            names.append(nombre)
    return "\n".join(names) if names else fallback


def write_peru_bot_processes_workbook(processes: list[dict], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Procesos Peru"
    _append_row(ws, PERU_BOT_HEADERS)
    _style_header(ws[1], BRAND_FOREST)
    for proc in processes:
        detail = proc.get("detail") or {}
        reporte = detail.get("reporte") or {}
        _append_row(ws, [
            proc.get("radicado", ""), proc.get("client_name", ""),
            detail.get("despacho") or proc.get("organo", ""),
            *[reporte.get(key, "") for key in PERU_BOT_REPORT_KEYS],
            _grouped_actor_names(detail, "DEMANDANTE", proc.get("demandante", "")),
            _grouped_actor_names(detail, "DEMANDADO", proc.get("demandado", "")),
            detail.get("valorParte", ""),
            *[detail.get(key, "") for key in PERU_BOT_DOCUMENTO_KEYS],
            detail.get("error", ""),
        ])
    for idx, width in enumerate(PERU_BOT_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "top" if cell.row > 1 else "center"
            cell.alignment = alignment
            if cell.row > 1:
                font = copy(cell.font)
                font.name = "Calibri"
                cell.font = font
                if cell.row % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=BRAND_MIST)
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
